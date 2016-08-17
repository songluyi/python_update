# -*- coding: utf-8 -*-
# 2016/8/16 14:32
"""
-------------------------------------------------------------------------------
Function:   自动将excel 数据通过selenium 模拟的方式填入表单
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""

import os,time
from openpyxl import load_workbook
from selenium import webdriver
def excel_data():
    file_name='活动周期设置.xlsx'
    wb_load = load_workbook(filename=file_name)
    sheets = wb_load.get_sheet_names()
    ws_load = wb_load.get_sheet_by_name(sheets[0])
    header_name=['序号', '阶段', '简称', '名称','摘要']
    full_shit_data=[]
    ID=int(input('请输入导入的序号:'))#设置导入序号
    row_line=int(input('请输入excel数据所在行数:'))
    for row in range(row_line,500):#设置excel导入行数
        row_data=[]
        if ws_load.cell(row=row, column=1).value is None and ws_load.cell(row=row, column=2).value is not None:
            row_data.append(ID)
            row_data.append(ws_load.cell(row=row, column=2).value)
            row_data.append(ws_load.cell(row=row, column=3).value)
            row_data.append(ws_load.cell(row=row, column=4).value)
            row_data.append(ws_load.cell(row=row, column=5).value)
            make_dict=dict(zip(header_name,row_data))
            full_shit_data.append(make_dict)
            ID+=1
    return full_shit_data

if __name__ == '__main__':
    data_list=excel_data()
    print(data_list)
    chromedriver = "C:\\Users\\Administrator\\AppData\\Local\\Google\\Chrome\\Application\\chromedriver.exe"#chrome所在驱动
    os.environ["webdriver.chrome.driver"] = chromedriver
    driver = webdriver.Chrome(chromedriver)
    driver.get("http://ebsapptest.nerin.com:8010/OA_HTML/RF.jsp?function_id=16402&resp_id=22593&resp_appl_id=275&security_group_id=0&lang_code=ZHS&params=f9uD-Csf0yJK45k-D90P04yPVdyR11TSVV5x7p0BhzA")
    driver.find_element_by_id('usernameField').send_keys('cg02')#登陆用户名
    driver.find_element_by_id('passwordField').send_keys('1q2w3e4r5t')#登陆密码
    driver.find_element_by_id('SubmitButton').click()
    driver.find_element_by_xpath('//*[@id="N3:UpdateImageEnabled:0"]/img').click()
    driver.find_element_by_xpath('//*[@id="LifecycleDetailsPhases"]/table[2]/tbody/tr[7]/th/button').click()
    count=0
    for i in data_list:
        print(i)
        print(count)
        print(i['序号'])
        time.sleep(1)
        driver.find_element_by_xpath('//*[@id="N31:PhaseSequence:%s"]'%count).send_keys(i['序号'])
        print(i['阶段'])
        driver.find_element_by_xpath('//*[@id="N31:StatusName:%s"]'%count).send_keys(i['阶段'])
        print(i['简称'])
        driver.find_element_by_xpath('//*[@id="N31:PhaseShortName:%s"]'%count).send_keys(i['简称'])
        print(i['名称'])
        time.sleep(3)
        element=driver.find_element_by_xpath('//*[@id="N31:PhaseName:%s"]'%count)
        element.clear()
        element.send_keys(i['名称'])
        print(i['摘要'])
        time.sleep(1)
        element=driver.find_element_by_xpath('//*[@id="N31:PhaseDescription:%s"]'%count)
        element.clear()
        element.send_keys(i['摘要'])
        if count==4:
            driver.find_element_by_xpath('//*[@id="LifecycleDetailsPhases"]/table[2]/tbody/tr[7]/th/button').click()
            count=0
        else:
            count=count+1
        print(count)
    # time.sleep(100)
    # driver.close()
    # driver.quit()