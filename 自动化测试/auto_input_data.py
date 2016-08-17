# -*- coding: utf-8 -*-
# 2016/8/16 13:18
"""
-------------------------------------------------------------------------------
Function:   自动将excel 数据通过selenium 模拟的方式填入表单
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com 
 
-------------------------------------------------------------------------------
"""

from selenium import webdriver
from openpyxl import load_workbook
def excel_data():
    file_name='活动周期设置.xlsx'
    wb_load = load_workbook(filename=file_name)
    sheets = wb_load.get_sheet_names()
    ws_load = wb_load.get_sheet_by_name(sheets[0])
    header_name=['序号', '阶段', '简称', '名称','摘要']
    full_shit_data=[]
    ID=int(input('请输入导入的序号:'))#设置导入序号
    row_line=int(input('请输入excel数据所在行数:'))
    for row in range(row_line,500):#设置excel导入行数
        row_data=[]
        if ws_load.cell(row=row, column=1).value is None and ws_load.cell(row=row, column=2).value is not None:
            row_data.append(ID)
            row_data.append(ws_load.cell(row=row, column=2).value)
            row_data.append(ws_load.cell(row=row, column=3).value)
            row_data.append(ws_load.cell(row=row, column=4).value)
            row_data.append(ws_load.cell(row=row, column=5).value)
            make_dict=dict(zip(header_name,row_data))
            full_shit_data.append(make_dict)
            ID+=1
    return full_shit_data
s=excel_data()
for i in s:
    print(i['序号'],i['阶段'],i['简称'],i['名称'],i['摘要'])




